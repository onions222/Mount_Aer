import xlrd
import  xlwt
import numpy as np
import matplotlib.pyplot as plt
import math
from numpy import polyfit, poly1d
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

pix_high = 2028
stick = 130
brick = 50

def cal_line(x1, y1, x2, y2):
    A = y1 - y2
    B = x2 - x1
    C = (x1 * y2) - (x2 * y1)
    return A, B, C

def cal_cross_point(A1, B1, C1, A2, B2, C2):
    #cal the intersection point o
    inter_x = (B1 * C2 - C1 * B2) / (A1 * B2 - B1 * A2)
    inter_y = (A1 * C2 - C1 * A2) / (A2 * B1 - B2 * A1)
    return inter_x, inter_y

def cal_distance(a, b, c, d):
    AC = math.sqrt(pow(a - c, 2) + pow(b - d, 2))
    return AC

def cal_medium_point(listx, listy):
    #the ordinates of A B C D
    x_A, x_B, x_C, x_D= listx[0], listx[1], listx[2], listx[3]
    y_A, y_B, y_C, y_D = listy[0], listy[1], listy[2], listy[3]

    # cal line AD
    A1 = y_D - y_A
    B1 = x_A - x_D
    C1 = (x_D * y_A) - (x_A * y_D)

    #cal line BC
    A2 = y_C - y_B
    B2 = x_B - x_C
    C2 = (x_C * y_B) - (x_B * y_C)

    #cal the intersection point o
    inter_x = (B1 * C2 - C1 * B2) / (A1 * B2 - B1 * A2)
    inter_y = (A1 * C2 - C1 * A2) / (A2 * B1 - B2 * A1)

    return inter_x, inter_y

def read_excel(filename, index):
    datafile = filename
    wb = xlrd.open_workbook(filename = datafile)

    #按index找到所需的sheet
    work_sheet = wb.sheet_by_index(index)

    #读取x和y的值
    x_col = work_sheet.col_values(1)
    y_col = work_sheet.col_values(2)
    x_list = np.array(x_col)
    y_list = np.array(y_col)

    return x_list, y_list

def cal_cross_ratio(x_list_process, y_list_process, listx1, listy1, listx2, listy2):
    '''
    :param x_list_process:
    :param y_list_process:
    :param listx1:  GHIJ
    :param listy1:  GHIJ
    :param listx2:  ABCD
    :param listy2:  ABCD
    :return: 高度和远度
    '''
    x_list = x_list_process
    y_list = y_list_process

    #删除第一个字母x y
    listx = np.delete(x_list, 0)
    listy = np.delete(y_list, 0)

    #计算竖直部分 点和线
    Kx, Ky = cal_medium_point(listx1, listy1)
    Kx, Ky = float(Kx), float(Ky)
    A1, B1, C1 = cal_line(listx1[0], listy1[0], listx1[2], listy1[2])      # line GI
    A4, B4, C4 = A1, B1, (-A1 * Kx - B1 * Ky)                              # line LK
    A2, B2, C2 = cal_line(listx1[0], listy1[0], listx1[1], listy1[1])      # line GH
    A3, B3, C3 = cal_line(listx1[2], listy1[2], listx1[3], listy1[3])      # line IJ
    Lx, Ly = cal_cross_point(A2, B2, C2, A4, B4, C4)                       # point L
    Mx, My = cal_cross_point(A3, B3, C3, A4, B4, C4)                       # point M

    #计算水平方向  点和线
    Qx, Qy = cal_medium_point(listx2, listy2)
    Qx, Qy = float(Qx), float(Qy)

    A5, B5, C5 = cal_line(listx2[0], listy2[0], listx2[2], listy2[2])      # line AC
    A6, B6, C6 = cal_line(listx2[0], listy2[0], listx2[1], listy2[1])      # line AB
    A7, B7, C7 = cal_line(listx2[1], listy2[1], listx2[3], listy2[3])      # line BD
    A8, B8, C8 = A6, B6, (-A6 * Qx - B6 * Qy)                              # line RP
    A9, B9, C9 = cal_line(listx2[2], listy2[2], listx2[3], listy2[3])      # line CD

    #无穷远点
    infx, infy = cal_cross_point(A6, B6, C6, A9, B9, C9)
    A10, B10, C10 = cal_line(Lx, Ly, infx, infy)                           # line infL

    list_real_x = []
    list_real_y = []

    #处理像素坐标转化高度远度
    for i in range(len(listx)):
        itex = float(listx[i])
        itey = float(listy[i])

        A11, B11, C11 = A1, B1, (-A1 * itex -B1 * itey)     #line OS
        Wx, Wy = cal_cross_point(A11, B11, C11, A10, B10, C10)    #point W

        #   distcance of W and L
        WL = cal_distance(Wx, Wy, Lx, Ly)
        h = WL
        infL = cal_distance(infx, infy, Lx, Ly)
        l = infL

        # MK = cal_distance(Mx, My, Kx, Ky)
        MK = abs(Ky - My)
        # ML = cal_distance(Mx, My, Lx, Ly)
        ML = abs(Ly - My)
        # KL = cal_distance(Kx,Ky,Lx,Ly)
        KL = abs(Ly - Ky)

        '''
        ================================= cal y direction====================================
        '''

        if Ky < itey < My:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                OK = abs(itey - Ky)
                alpha = (OL * MK) / (OK * ML)
                y = (stick * alpha) / (2 * stick - 1)
                real_y = -y
            else:
                OK = abs(itey - Ky)
                OL = abs(itey - Ly)
                alpha = (MK * OL) / (OK * ML)
                y = (stick / 2) / (2 * alpha - 1)
                real_y = - stick / 2 - y

        if Ly < itey < Ky:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                OM = abs(itey - My)
                OK = abs(KL - OL)
                alpha = (KL * OM) / (OK * ML)
                y = (stick * alpha - stick) / (2 * alpha - 1)
                real_y = -y
            else:
                OM = abs(itey - My)
                OK = abs(itey - Ky)
                alpha = (OM * KL) / (OK * ML)
                y = (stick / 2) / (2 * alpha - 1)
                real_y = -stick / 2 + y

        if itey < Ly:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                OK = OL + KL
                OM = OL + ML
                alpha = (OK * ML) / (KL * OM)
                y = (130 * alpha - 130) / (2 - alpha)
                real_y = y
            else:
                OM = abs(itey - My)
                OK = abs(itey - Ky)
                alpha = (ML * OK) / (KL * OM)
                y = (stick - stick * alpha) / (alpha - 2)
                real_y = y

        if itey > My:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                OK = abs(OL - KL)
                alpha = (ML * OK) / (MK * OL)
                y = stick / (2 - alpha)
                real_y = -y
            else:
                OK = abs(itey - Ky)
                OL = abs(itey - Ly)
                alpha = (OK * ML) / (MK * OL)
                y = (stick - stick * alpha) / (alpha - 2)
                real_y = -stick - y

        if itey == Ly:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                real_y = - OL
            else:
                real_y = 0
            # real_y = 0
        if itey == Ky:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                real_y = - OL
            else:
                real_y = - stick / 2
            # real_y = - stick / 2
        if itey == My:
            if itex <= Lx:
                OW = cal_distance(itex, itey, Wx, Wy)
                OL = (l * OW) / (h + l)
                real_y = - OL
            else:
                real_y = -stick
            # real_y = - stick
        list_real_y.append(real_y)

        '''
        ===================================== cal x direction=================================
        '''

        Rx, Ry = cal_cross_point(A5, B5, C5, A8, B8, C8)
        Px, Py = cal_cross_point(A8, B8, C8, A7, B7, C7)
        RQ = cal_distance(Rx, Ry, Qx, Qy)
        RP = cal_distance(Px, Py, Rx, Ry)
        PQ = cal_distance(Px, Py, Qx, Qy)

        Sx, Sy = cal_cross_point(A5, B5, C5, A6, B6, C6)    # point S

        if itex > Px:
            RS = cal_distance(Rx, Ry, Sx, Sy)
            PS = cal_distance(Px, Py, itex, itey)
            beta = (RQ * PS) / (PQ * RS)
            x = ((brick / 2) * 3 * beta - (brick / 2) * 3) / (3 - 2 * beta)
            real_x = - brick - x

        if Px < itex < Qx:
            RS = cal_distance(Rx, Ry, Sx, Sy)
            PS = cal_distance(Px, Py, Sx, Sy)
            beta = (RS * PQ) / (PS * RQ)
            x = (brick / 2) / (3 * beta - 2)
            real_x = - brick / 2 - x

        if Rx < itex < Px:
            PS = cal_distance(Px, Py, Sx, Sy)
            QS = cal_distance(Qx, Qy, Sx, Sy)
            beta = (RP * QS) / (PS * RQ)
            x = (brick  * beta - brick) / (3 * beta - 1)
            real_x = -x

        if itex < Rx:
            PS = cal_distance(Px, Py, Sx, Sy)
            QS = cal_distance(Qx, Qy, Sx, Sy)
            beta = (PS * RQ) / (RP * QS)
            x = (brick * beta - brick) / (3 - beta)
            real_x = x

        if itex == Qx:
            x = brick
            real_x = -x

        if itex == Px:
            x = brick / 2
            real_x = -x

        if itex == Rx:
            real_x = 0
        list_real_x.append(real_x)

    return list_real_x, list_real_y

if __name__ == '__main__':
    datafile = r'bqz0203l03_final.xlsx'
    listx1 = np.array([2058, 2396, 2122, 2486])
    listy1 = np.array([1402, 1356, 1446, 1388])
    listx2 = np.array([1534, 1560, 1536, 1559])
    listy2 = np.array([1153, 1175, 1466, 1529])

    wb = load_workbook(datafile)
    wb_real = Workbook()
    sheet_list = wb.sheetnames
    for index in range(0, len(sheet_list)):
        list_excel_x, list_excel_y = read_excel(datafile, index)
        list_real_x, list_real_y = np.array(cal_cross_ratio(list_excel_x, list_excel_y, listx1, listy1, listx2, listy2))
        sheetname = 'sheet' + str(index)
        wb_real.create_sheet(sheetname)
        ws = wb_real['sheet' + str(index)]
        for i in range(1, len(list_real_x)):
            ws['A' + str(i)].value = list_real_x[i]
            ws['B' + str(i)].value = list_real_y[i]
    wb_real.save('bqz0203l.xlsx')













